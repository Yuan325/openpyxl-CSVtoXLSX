from openpyxl import Workbook
import csv
from datetime import datetime
date = datetime.today().strftime('%Y%m%d')

output = 'output/'+date+'_SQLoutput.xlsx'
final = 'output/final_SQLoutput.xlsx'
wb = Workbook()
ws = wb.active
with open('output/'+date+'_SQLoutput.csv', 'r') as f:
    for row in csv.reader(f):
        ws.append(row)
wb.save(output)

import openpyxl
ss = openpyxl.load_workbook(output)
ss_sheet = ss['Sheet']
ss_sheet.title = 'Summary'
ss.save(output)


wb1 = openpyxl.load_workbook(output)
wb1_s = wb1['Summary']
wb2 = Workbook()
wb2_s = wb2.create_sheet("Summary", 0)
wb2_s2 = wb2.create_sheet("DB2", 1)
wb2_s3 = wb2.create_sheet("DB3",2)
wb2_s4 = wb2.create_sheet("DB4",3)
wb2_s5 = wb2.create_sheet("DB5",4)
wb2_s6 = wb2.create_sheet("DB6",5)
wb2_s7 = wb2.create_sheet("DB7",6)

for row in wb1_s['A1':'B8']:
    for cell in row:
        wb2_s[cell.coordinate].value = cell.value

def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)

    return rangeSelected

def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

db2_tot = int(wb1_s.cell(3,2).value)
db3_tot = int(wb1_s.cell(4,2).value)
db4_tot = int(wb1_s.cell(5,2).value)
db5_tot = int(wb1_s.cell(6,2).value)
db6_tot = int(wb1_s.cell(7,2).value)
db7_tot = int(wb1_s.cell(8,2).value)
db2_start = 13 + 1
db2_end = db2_start + 2 + db2_tot
db3_start = db2_end + 5
db3_end = db3_start+2+db3_tot
db4_start = db3_end+5
db4_end = db4_start+2+db4_tot
db5_start = db4_end+5
db5_end = db5_start+2+db5_tot
db6_start = db5_end+5
db6_end = db6_start+2+db6_tot
db7_start =db6_end+5
db7_end = db7_start+2+db7_tot

print("processing copy and paste")
selectedRange = copyRange(1,db2_start,18,db2_end,wb1_s)
pastingRagne = pasteRange(1,1,18,3+db2_tot,wb2_s2,selectedRange)
selectedRange = copyRange(1,db3_start,18,db3_end,wb1_s)
pastingRagne = pasteRange(1,1,18,3+db3_tot,wb2_s3,selectedRange)
selectedRange = copyRange(1,db4_start,18,db4_end,wb1_s)
pastingRagne = pasteRange(1,1,18,3+db4_tot,wb2_s4,selectedRange)
selectedRange = copyRange(1,db5_start,18,db5_end,wb1_s)
pastingRagne = pasteRange(1,1,18,3+db5_tot,wb2_s5,selectedRange)
selectedRange = copyRange(1,db6_start,18,db6_end,wb1_s)
pastingRagne = pasteRange(1,1,18,3+db6_tot,wb2_s6,selectedRange)
selectedRange = copyRange(1,db7_start,18,db7_end,wb1_s)
pastingRagne = pasteRange(1,1,18,3+db7_tot,wb2_s7,selectedRange)
print ("Done processing")

wb2.save(final)
